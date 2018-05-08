#!/usr/bin/python

""" Author: Peter Swanson
            pswanson@ucdavis.edu

    Description: A class to make interacting with spreadsheets via the openpyxl library easier

    Version: Python 2.7
    Requirements: openpyxl """

from openpyxl import Workbook, load_workbook
import os

class Spreadsheet(object):
    """ Class to simplify operating on Excel sheets """

    def __init__(self, filename, load=True):
        """ @filename - The name of the excel file including the extension
            @load - Load the file if true, create it if false """

        self.filename = filename

        # If load is False, recreate excel file on every run
        # Else load the file
        if load is True:
            self.file = load_workbook(filename=filename)
        else:
            if os.path.isfile(filename):
                os.remove(filename)

            self.file = Workbook()
            self.save()

        self.sheet = self.file.active

    def save(self):
        """ Save the file """

        self.file.save(filename=self.filename)

    def write(self, content, cell):
        """ Write to a cell
            @content - The content to write
            @cell - The cell to write to """

        self.sheet[cell] = content
        self.save()

    def read_column(self, col, header=True, start_row=1):
        """ Read values from a column
            @col - The column to read
            @header - True if the column has a header """

        # Column stored in a dict with a single list
        values = []
        column = {}

        row = start_row

        # Header is dict key, values is a list of cell values from the column
        if header:
            column[self.sheet[row][col - 1].value] = values
            row += 1
        else:
            column["None"] = values

        # Store values
        while True:
            cell = self.sheet.cell(row=row, column=col).value

            if cell is None:
                break

            values.append(cell)
            row += 1

        return column

    def read_row(self, row, start_col=1):
        """ Read values from a row
            @row - The row to read
            @start_col - The column to read past"""

        # Row stored in a dict with a single list
        values = []
        obj = {}
        obj['Row: ' + str(row)] = values

        col = start_col

        # Store values
        while True:
            cell = self.sheet.cell(row=row, column=col).value

            if cell is None:
                break

            values.append(cell)
            col += 1

        return obj

    def write_column(self, col, content, start_row=1, bold=False, italics=False):
        """ Write values to a column
            @col - The column to read
            @start_row - The row to start writing at
            @content - a list of values to write """

        # Column stored in a dict with a single list
        row = start_row

        # Store values
        for value in content:
            cell = self.sheet.cell(row=row, column=col, value=value.encode('unicode_escape').decode('utf-8'))

            cell.font = cell.font.copy(bold=bold, italic=italics)
            row += 1

        self.save()

    def write_row(self, row, content, start_col=1, bold=False, italics=False):
        """ Write values to a row
            @row - The row to write to
            @start_col - The column to start writing at
            @content - a list of values to write """

        # Column stored in a dict with a single list
        col = start_col

        # Store values
        for value in content:
            cell = self.sheet.cell(row=row, column=col, value=value.encode('unicode_escape').decode('utf-8'))
            cell.font = cell.font.copy(bold=bold, italic=italics)
            col += 1

        self.save()

    def create_sheets(self, values):
        """ Create sheets """

        for value in values:
            self.file.create_sheet(value)

        self.save()