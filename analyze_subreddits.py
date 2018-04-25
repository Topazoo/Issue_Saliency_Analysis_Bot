from openpyxl import Workbook, load_workbook
import os

class Subreddit(object):
    """ Class to hold subreddit information """

    def __init__(self, row):
        self.name = row

class Spreadsheet(object):
    """ Class to simplify operating on Excel sheets """

    def __init__(self, filename, load=True):
        self.filename = filename

        # If load is False, recreate excel file on every run
        # Else load the file
        if load is True:
            self.file = load_workbook(filename=filename)
        else:
            if os.path.isfile(filename):
                os.remove(filename)

            self.file = Workbook()
            self.save()

        self.sheet = self.file.active

    def save(self):
        """ Save the file """

        self.file.save(filename=self.filename)

    def write(self, content, cell):
        """ Write to a cell """

        self.sheet[cell] = content
        self.save()

class Bot(object):
    """ Class to contain highest-order program operations """

    def __init__(self):
        # Open spreadsheets for inputs and results
        self.input_sheet = Spreadsheet("subreddits.xlsx")
        self.output_file = Spreadsheet("results.xlsx", False)

        # Create a list of subreddits
        self.subreddits = []

def main():
    bot = Bot()
    bot.output_file.write("Test", "B1")

if __name__ == '__main__':
    main()