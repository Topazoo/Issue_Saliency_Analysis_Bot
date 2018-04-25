#!/usr/bin/python

""" Author: Peter Swanson
            pswanson@ucdavis.edu

    Description: A class to make interacting with spreadsheets via the openpyxl library easier

    Version: Python 2.7
    Requirements: openpyxl """

from openpyxl import Workbook, load_workbook
import os

class Spreadsheet(object):
    """ Class to simplify operating on Excel sheets """

    def __init__(self, filename, load=True):
        """ Params:
            @filename - The name of the excel file including the extension
            @load - Load the file if true, create it if false """

        self.filename = filename

        # If load is False, recreate excel file on every run
        # Else load the file
        if load is True:
            self.file = load_workbook(filename=filename)
        else:
            if os.path.isfile(filename):
                os.remove(filename)

            self.file = Workbook()
            self.save()

        self.sheet = self.file.active

    def save(self):
        """ Save the file """

        self.file.save(filename=self.filename)

    def write(self, content, cell):
        """ Write to a cell
            Params:
            @content - The content to write
            @cell - The cell to write to """

        self.sheet[cell] = content
        self.save()

    def read_column(self, col, header=True):
        """ Read values from a column
            Params:
            @col - The column to read
            @header - True if the column has a header """

        # Column stored in a dict with a single list
        values = []
        column = {}

        row = 1

        # Header is dict key, values is a list of cell values from the column
        if header:
            column[self.sheet[row][col].value] = values
            row += 1
        else:
            column["None"] = values

        # Store values
        while True:
            cell = self.sheet[row][col].value

            if cell is None:
                break

            values.append(cell)
            row += 1

        return column
